import streamlit as st
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd
import io
import re
import zipfile
import requests
import time
import plotly.express as px
from fpdf import FPDF
from datetime import datetime, timedelta

# --- BRANDING & CONSTANTS ---
COLOR_PRIMARY = (16, 43, 85) 
COLOR_SECONDARY = (212, 175, 55) 
COLOR_TEXT = (50, 50, 50) 

GOOGLE_SHEET_URL = "https://docs.google.com/spreadsheets/d/14ykcy9qOUPu-wLp7Xzp6SJWqVmXUzGAT/export?format=xlsx"

class AmaniInvoice(FPDF):
    def header(self):
        self.set_font("Helvetica", "B", 22)
        self.set_text_color(*COLOR_PRIMARY)
        self.cell(100, 10, "AMANI DAIRIES", ln=0)
        self.set_font("Helvetica", "I", 10)
        self.set_text_color(100, 100, 100)
        self.cell(90, 10, "Reliable - Fresh - Local", ln=1, align="R")
        self.set_draw_color(*COLOR_SECONDARY)
        self.set_line_width(0.8)
        self.line(10, 20, 200, 20)
        self.ln(6)

    def draw_calendar_grid(self, daily_data, billed_month_str):
        try:
            year_match = re.search(r'20\d{2}', billed_month_str)
            y = int(year_match.group(0)) if year_match else datetime.now().year
            base_date = datetime(y, 1, 1)
            months = ["jan","feb","mar","apr","may","jun","jul","aug","sep","oct","nov","dec"]
            for i, m in enumerate(months):
                if m in billed_month_str.lower():
                    base_date = datetime(y, i+1, 1)
                    break
        except:
            base_date = datetime.now()
        
        self.set_font("Helvetica", "B", 8)
        self.set_text_color(*COLOR_PRIMARY)
        self.cell(0, 6, "DAILY CONSUMPTION BREAKDOWN", ln=1)
        w_small = 5.8; h = 5 
        self.set_fill_color(245, 245, 245); self.set_text_color(0,0,0)
        self.cell(20, h, "Date", 1, 0, 'C', True)
        for d in range(1, 32): self.cell(w_small, h, str(d), 1, 0, 'C', True)
        self.ln(h)
        self.cell(20, h, "Day", 1, 0, 'C')
        self.set_font("Helvetica", "", 6)
        for d in range(1, 32):
            try: day_str = (base_date + timedelta(days=d-1)).strftime("%a")
            except: day_str = ""
            self.cell(w_small, h, day_str, 1, 0, 'C')
        self.ln(h)
        self.set_font("Helvetica", "B", 7)
        self.cell(20, h, "Litres", 1, 0, 'C')
        for d in range(1, 32):
            val = daily_data.get(d, 0)
            self.cell(w_small, h, str(int(val)) if val > 0 else "-", 1, 0, 'C')
        self.ln(h + 6)

def create_branded_pdf(cust_all_data, billed_month_str):
    def make_safe(t):
        t = str(t).replace('\u2022', '-').replace('\u2013', '-').replace('\u2014', '-').replace('\u2219', '.')
        return re.sub(r'[^\x00-\x7f]', r'', t).encode('latin-1', 'replace').decode('latin-1')
    def fmt(val): return f"{val:,.2f}"
    pdf = AmaniInvoice(orientation='P', unit='mm', format='A4')
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 9); pdf.set_text_color(*COLOR_PRIMARY)
    pdf.cell(100, 5, "BILL TO:"); pdf.ln(1)
    pdf.set_font("Helvetica", "B", 13); pdf.set_text_color(*COLOR_TEXT)
    pdf.cell(100, 7, make_safe(cust_all_data['name']), ln=1)
    pdf.set_xy(140, 26); pdf.set_font("Helvetica", "B", 9); pdf.set_text_color(*COLOR_PRIMARY)
    pdf.cell(50, 5, "BILLING PERIOD:", ln=1, align="R")
    pdf.set_xy(140, 31); pdf.set_font("Helvetica", "B", 12); pdf.set_text_color(*COLOR_TEXT)
    pdf.cell(50, 7, make_safe(billed_month_str).upper(), ln=1, align="R")
    pdf.set_xy(135, 42); pdf.set_fill_color(255, 255, 255); pdf.set_draw_color(0,0,0); pdf.set_text_color(200, 0, 0); pdf.set_font("Helvetica", "B", 11)
    pdf.cell(60, 8, "BALANCE DUE", 1, 1, "C")
    pdf.set_xy(10, 55); pdf.draw_calendar_grid(cust_all_data['daily_liters'], billed_month_str)
    pdf.set_fill_color(*COLOR_PRIMARY); pdf.set_text_color(255, 255, 255); pdf.set_font("Helvetica", "B", 9)
    pdf.cell(80, 8, "  Description", 1, 0, "L", True); pdf.cell(30, 8, "Total Qty (L)", 1, 0, "C", True); pdf.cell(30, 8, "Rate (KES)", 1, 0, "C", True); pdf.cell(50, 8, "Total (KES)  ", 1, 1, "R", True)
    pdf.set_text_color(*COLOR_TEXT); pdf.set_font("Helvetica", "", 10)
    pdf.cell(80, 10, "  Fresh Milk Supplied", 1, 0, "L"); pdf.cell(30, 10, f"{cust_all_data['billed_qty']:.1f}", 1, 0, "C"); pdf.cell(30, 10, f"{fmt(cust_all_data['rate'])}", 1, 0, "C")
    pdf.set_font("Helvetica", "B", 10); pdf.cell(50, 10, f"{fmt(cust_all_data['total_bill'])}  ", 1, 1, "R")
    pdf.ln(1); pdf.set_x(120); pdf.set_font("Helvetica", "", 9)
    pdf.cell(40, 6, "Sub-Total:", 0, 0, "R"); pdf.cell(40, 6, f"{fmt(cust_all_data['total_bill'])}", 0, 1, "R")
    pdf.set_x(120); pdf.cell(40, 6, "Pre-Paid:", 0, 0, "R"); pdf.cell(40, 6, f"- {fmt(cust_all_data['prepaid'])}", 0, 1, "R")
    pdf.set_x(120); pdf.set_font("Helvetica", "B", 11); pdf.set_text_color(*COLOR_PRIMARY)
    pdf.cell(40, 10, "TOTAL DUE:", "T", 0, "R"); pdf.cell(40, 10, f"KES {fmt(cust_all_data['balance'])}", "T", 1, "R")
    if cust_all_data['spoilt_qty'] > 0:
        pdf.ln(8); pdf.set_font("Helvetica", "B", 8); pdf.set_text_color(200, 0, 0)
        pdf.cell(0, 5, "SPOILT MILK NOTICE (Excluded from Total Bill):", ln=1)
        pdf.set_font("Helvetica", "", 8); pdf.set_text_color(*COLOR_TEXT)
        spoilt_str = ", ".join([f"Day {d} ({q}L)" for d, q in cust_all_data['spoilt_details']])
        pdf.cell(0, 4, f"The following recorded milk was spoilt: {spoilt_str}. Total: {cust_all_data['spoilt_qty']:.1f}L.", ln=1)
    pdf.ln(15); current_y = pdf.get_y(); pdf.set_fill_color(252, 248, 227); pdf.set_draw_color(*COLOR_SECONDARY); pdf.rect(10, current_y, 190, 25, 'FD')
    pdf.set_y(current_y + 3); pdf.set_font("Helvetica", "B", 9); pdf.set_text_color(*COLOR_PRIMARY); pdf.cell(190, 4, "PAYMENT METHODS", ln=1, align="C")
    pdf.set_font("Helvetica", "B", 12); pdf.cell(190, 6, "M-PESA POCHI LA BIASHARA", ln=1, align="C")
    pdf.set_font("Helvetica", "B", 18); pdf.cell(190, 8, "0722 686 720", ln=1, align="C")
    return pdf.output()

def clean_num(value):
    if value is None or str(value).strip() in ["", "-", "None"]: return 0.0
    try: return float(value)
    except: return 0.0

def get_month_data(ws):
    all_data = []
    if not isinstance(ws, Worksheet): return []
    if not ws.cell(row=2, column=10).value: return []

    for col in range(10, ws.max_column + 1):
        name = ws.cell(row=2, column=col).value
        if not name or any(skip in str(name) for skip in ["Total", "Unaccounted", "Summary", "Fridge"]): break
        rate = clean_num(ws.cell(row=3, column=col).value)
        prepaid = clean_num(ws.cell(row=37, column=col).value)
        total_qty = 0; spoilt_qty = 0; daily_dict = {}; spoilt_list = []
        for row in range(4, 35):
            cell = ws.cell(row=row, column=col); val = clean_num(cell.value)
            day = row - 3; daily_dict[day] = val
            fill = str(cell.fill.start_color.index)
            if fill in ['FFFF0000', '2'] and val > 0:
                spoilt_qty += val; spoilt_list.append((day, val))
            else: total_qty += val
        all_data.append({
            "name": name, "billed_qty": total_qty, "spoilt_qty": spoilt_qty,
            "rate": rate, "total_bill": total_qty * rate, "lost_revenue": spoilt_qty * rate,
            "prepaid": prepaid, "balance": (total_qty * rate) - prepaid,
            "daily_liters": daily_dict, "spoilt_details": spoilt_list
        })
    return all_data

# --- APP INTERFACE ---
st.set_page_config(page_title="Amani Dairies Dashboard", layout="wide")
st.title("🥛 Amani Dairies Performance Tracker")

c1, c2 = st.columns([1, 1])
with c1:
    if st.button("🔄 Sync with Google Sheet", use_container_width=True):
        try:
            with st.spinner("Fetching data from Cloud..."):
                session = requests.Session()
                session.mount('https://', requests.adapters.HTTPAdapter(max_retries=3))
                # Cache-busting ensures you don't get an old version of the file
                response = session.get(f"{GOOGLE_SHEET_URL}&t={int(time.time())}", timeout=60)
                if response.status_code == 200:
                    st.session_state['data_file'] = io.BytesIO(response.content)
                    st.success("Cloud Sync Successful!")
                else: st.error("Failed to reach Google Sheet.")
        except Exception as e: st.error(f"Sync Error: {e}")

with c2:
    uploaded_file = st.file_uploader("Or Upload Manual Excel File", type=["xlsx", "xlsm"])
    if uploaded_file: st.session_state['data_file'] = uploaded_file

if 'data_file' in st.session_state:
    wb = openpyxl.load_workbook(st.session_state['data_file'], data_only=True)
    
    # WE REMOVED THE NUMERIC FILTER HERE TO SHOW ALL MONTHS
    available_sheets = [s for s in wb.sheetnames if not any(x in s.lower() for x in ["summary", "data", "client", "total", "template"])]
    
    with st.spinner("Calculating Business Stats..."):
        all_months_results = {sheet: get_month_data(wb[sheet]) for sheet in available_sheets if get_month_data(wb[sheet])}

    if not all_months_results:
        st.warning("No valid billing sheets found. Check Column J.")
    else:
        # --- BUSINESS OVERVIEW SECTION ---
        st.header("📊 Cumulative Business Overview")
        total_rev = sum(sum(c['total_bill'] for c in data) for data in all_months_results.values())
        total_loss = sum(sum(c['lost_revenue'] for c in data) for data in all_months_results.values())
        total_liters = sum(sum(c['billed_qty'] for c in data) for data in all_months_results.values())
        
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Lifetime Revenue", f"KES {total_rev:,.0f}")
        m2.metric("Revenue Lost (Spoilage)", f"KES {total_loss:,.0f}", delta=f"{(total_loss/(total_rev+1)*100):.1f}% Loss", delta_color="inverse")
        m3.metric("Liters Supplied", f"{total_liters:,.1f} L")
        m4.metric("Active Billing Months", len(all_months_results))

        st.divider()
        
        # --- VISUALIZATIONS SECTION ---
        chart_data = [{"Month": m, "Revenue": sum(c['total_bill'] for c in data), "Loss": sum(c['lost_revenue'] for c in data)} for m, data in all_months_results.items()]
        df_trends = pd.DataFrame(chart_data)
        c_left, c_right = st.columns(2)
        with c_left: st.plotly_chart(px.bar(df_trends, x="Month", y=["Revenue", "Loss"], barmode="group", title="Revenue vs. Spoilage Loss", color_discrete_map={"Revenue": "#102B55", "Loss": "#C80000"}), use_container_width=True)
        with c_right:
            cust_totals = {}
            for m_data in all_months_results.values():
                for c in m_data: cust_totals[c['name']] = cust_totals.get(c['name'], 0) + c['total_bill']
            df_cust = pd.DataFrame(list(cust_totals.items()), columns=['Customer', 'Revenue']).sort_values("Revenue", ascending=False).head(10)
            st.plotly_chart(px.pie(df_cust, values='Revenue', names='Customer', title="Top 10 High-Value Clients"), use_container_width=True)

        st.divider()
        
        # --- INVOICE & TABLE SECTION ---
        # Advanced Sort: Puts 2026/2025 at the top
        def sort_key(name):
            year = re.search(r'20\d{2}', name)
            y_val = int(year.group(0)) if year else 0
            m_map = {"jan":1,"feb":2,"mar":3,"apr":4,"may":5,"jun":6,"jul":7,"aug":8,"sep":9,"oct":10,"nov":11,"dec":12}
            m_val = 0
            for k, v in m_map.items():
                if k in name.lower(): m_val = v; break
            return (y_val, m_val)

        sorted_months = sorted(list(all_months_results.keys()), key=sort_key, reverse=True)
        
        target_month = st.selectbox("Select Month for Invoices:", sorted_months)
        month_df = pd.DataFrame(all_months_results[target_month])
        st.dataframe(month_df[['name', 'billed_qty', 'spoilt_qty', 'total_bill', 'prepaid', 'balance']], use_container_width=True)

        if st.button(f"📥 Download All Invoices for {target_month}"):
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, "w") as zf:
                for cust in all_months_results[target_month]:
                    zf.writestr(f"{cust['name']}.pdf", create_branded_pdf(cust, target_month))
            st.download_button("💾 Save ZIP File", zip_buffer.getvalue(), f"Amani_Invoices_{target_month}.zip")
